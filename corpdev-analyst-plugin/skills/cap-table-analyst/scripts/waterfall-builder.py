"""
Waterfall Builder
=================
Builds liquidation waterfall Excel models from structured cap table data.

Key mechanics:
- Convertibles/SAFEs: pari passu; supports accrued interest, CoC multiples,
  and implied-share conversion check
- Preferred: explicit seniority groups; pari passu within each group;
  per-class multiples; cumulative dividend accruals
- Conversion check: non-participating preferred always takes MAX(liq pref, as-converted).
  Iterative solve — classes that convert join the common pool and give back their T2 payout.
- Participating preferred: takes pref payout PLUS participates pro-rata in common residual
- Capped participating: participates up to a total return cap; iterative solve
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Color Palette ────────────────────────────────────────────────────────────
C = dict(
    nav='1F4E79',
    white='FFFFFF',
    blue_hdr='2E74B5',
    green_hdr='375623',
    gold_hdr='BF8F00',
    gray_hdr='595959',
    highlight_hdr='ED7D31',
    highlight_hdr2='C55A11',
    green_bg='E2EFDA',
    green_sub='C6EFCE',
    blue_bg='DDEEFF',
    blue_sub='BDD7EE',
    gold_bg='FFF2CC',
    gold_sub='FFE699',
    gray_bg='F2F2F2',
    highlight='FCE4D6',
    input_bg='EBF3FB',
    total_bg='D6DCE4',
    subtitle='D9E1F2',
    blue_input='0000FF',
    black='000000',
    gray_text='7F7F7F',
    converted_bg='F0F0F0',   # light gray for converted rows in T2
    converted_text='A0A0A0',
)

CURRENCY  = '$#,##0;($#,##0);"-"'
CURRENCY2 = '$#,##0.00;($#,##0.00);"-"'
NUMFMT    = '#,##0'
PCT       = '0.0%'
MULT      = '0.0"x"'


# ─── Cell / Row Helpers ───────────────────────────────────────────────────────

def h(ws, row, col, value=None, formula=None, bold=False, italic=False,
      fc='000000', bg=None, num_fmt=None, align='left', size=10,
      wrap=False, height=None):
    cell = ws.cell(row=row, column=col)
    cell.value = formula if formula else (value if value is not None else None)
    cell.font = Font(name='Arial', bold=bold, italic=italic, color=fc, size=size)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    if height:
        ws.row_dimensions[row].height = height
    return cell


def rh(ws, row, ht):
    ws.row_dimensions[row].height = ht


def section_header(ws, row, ncols, text, bg, fc='FFFFFF', height=22):
    ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
    h(ws, row, 1, text, bold=True, fc=fc, bg=bg, align='center', height=height)


def col_headers(ws, row, headers, bg, fc='FFFFFF', row_height=18):
    for col_idx, label in enumerate(headers, start=1):
        h(ws, row, col_idx, label, bold=True, fc=fc, bg=bg, align='center', wrap=True)
    rh(ws, row, row_height)


# ─── Input Normalization ──────────────────────────────────────────────────────

def _normalize_conv(c):
    """
    Normalize a convertible instrument entry.

    Accepted formats:
      Old tuple:  (name, face_value)
      New dict:
        {
          'name': str,
          'principal': float,          # face value / principal
          'accrued_interest': float,   # default 0 — adds to T1 payout claim
          'coc_multiple': float,       # default 1.0 — multiplied by principal on M&A
          'implied_shares': int,       # default 0 — if >0, triggers conversion check:
                                       #   instrument takes MAX(claim, implied_shares × common_pps)
        }

    Notes on accrued_interest vs coc_multiple:
      - accrued_interest: pre-computed dollar amount (e.g. for convertible notes with
        known outstanding interest). Adds to the base claim: claim = principal + accrued_interest.
      - coc_multiple: a contractual multiplier applied to principal on a change-of-control
        (some SAFEs and bridge notes have 1.5x or 2x M&A return provisions).
        claim = principal × coc_multiple + accrued_interest.
      Both can be used together.
    """
    if isinstance(c, (tuple, list)):
        return {
            'name': c[0], 'principal': float(c[1]),
            'accrued_interest': 0.0, 'coc_multiple': 1.0, 'implied_shares': 0,
        }
    d = dict(c)
    d.setdefault('accrued_interest', 0.0)
    d.setdefault('coc_multiple', 1.0)
    d.setdefault('implied_shares', 0)
    d['principal'] = float(d['principal'])
    return d


def _normalize_pref(c, default_multiple):
    """
    Normalize a preferred class entry.

    Accepted formats:
      Old tuple:  (name, shares, capital_invested)
                  (name, shares, capital_invested, pref_multiple)
      New dict:
        {
          'name': str,
          'shares': int,
          'capital_invested': float,

          # Liquidation preference
          'pref_multiple': float,          # default: global pref_multiple
          'cumulative_dividends': float,   # default 0 — accrued divs add to liq pref claim
                                           # liq_pref_claim = capital × multiple + cumulative_dividends

          # Seniority (for ordering T2 payouts)
          'seniority': int,                # default 1 — lower = more senior
                                           # classes with the SAME seniority number are pari passu
                                           # (pro-rata by claim within the group)

          # Participation structure (T3 behavior for non-converting classes)
          'participation': str,            # 'non-participating' | 'participating' | 'capped'
          'participation_cap': float|None  # for 'capped': total return cap multiple of capital
                                           # e.g. 3.0 → max total return = 3× capital_invested
        }

    Conversion logic (applied automatically in compute_waterfall):
      Non-participating preferred always takes MAX(liq_pref_payout, as_converted_value).
      If as-converted is higher, the class converts: it joins the common pool and gives
      back its T2 payout. This is computed iteratively and is NOT settable per-class —
      it is always the correct financial behavior.
    """
    if isinstance(c, (tuple, list)):
        return {
            'name': c[0], 'shares': int(c[1]), 'capital_invested': float(c[2]),
            'pref_multiple': c[3] if len(c) > 3 else default_multiple,
            'cumulative_dividends': 0.0,
            'seniority': 1,
            'participation': 'non-participating',
            'participation_cap': None,
        }
    d = dict(c)
    d.setdefault('pref_multiple', default_multiple)
    d.setdefault('cumulative_dividends', 0.0)
    d.setdefault('seniority', 1)
    d.setdefault('participation', 'non-participating')
    d.setdefault('participation_cap', None)
    return d


def _participation_label(pc):
    p = pc['participation']
    m = pc['pref_multiple']
    mult_str = f'{m:.2g}x'
    if p == 'non-participating':
        return f'{mult_str} Non-Participating'
    elif p == 'participating':
        return f'{mult_str} Fully Participating'
    elif p == 'capped':
        cap = pc.get('participation_cap')
        cap_str = f'{cap:.2g}x' if cap else '?x'
        return f'{mult_str} Participating (Cap: {cap_str})'
    return p


# ─── Waterfall Computation ────────────────────────────────────────────────────

def compute_waterfall(total_exit, conv_instruments_norm, pref_classes_norm, common_shares):
    """
    Compute the full liquidation waterfall.

    Returns a dict with all tier amounts, per-instrument/class payouts,
    conversion decisions, and per-share values. Used for:
      - Verification before writing the Excel model
      - Driving capped participation amounts (written as blue inputs in Excel)
      - Driving conversion check results (written as blue inputs in Excel)

    Algorithm
    ---------
    Tier 1 — Convertibles / SAFEs (always pari passu):
      claim_i = principal_i × coc_multiple_i + accrued_interest_i
      if sum(claims) ≤ remaining: each instrument paid in full
      else: pro-rata by claim
      Conversion check: if implied_shares_i > 0 and implied_shares_i × pps > claim_i,
        instrument receives equity value instead

    Tier 2 — Preferred (seniority groups, pari passu within group):
      Groups processed most-senior-first (lowest seniority number first).
      Within each group: pari passu — if group can't be fully covered,
        each class receives (claim_i / group_total) × remaining.
      liq_pref_claim_i = capital_invested_i × pref_multiple_i + cumulative_dividends_i

    Conversion check — non-participating preferred (iterative):
      A non-participating class converts if: shares × common_pps > liq_pref_payout.
      Iterative: converting classes join the common pool and free up their T2 payout,
      raising common_pps, potentially triggering more conversions. Repeat until stable.
      Participating preferred is NOT subject to this check (they already get pref + participation).

    Tier 3 — Common pool (+ participating + converted):
      Pool denominator = common_shares + converted_pref_shares + participating_pref_shares
      For capped participation: iterative solve (classes that hit their cap exit the pool).
    """

    # ── Tier 1 ────────────────────────────────────────────────────────────────
    t1 = []
    for c in conv_instruments_norm:
        claim = c['principal'] * c['coc_multiple'] + c['accrued_interest']
        t1.append({**c, 'claim': claim, 'payout': 0.0})

    t1_total_claims = sum(x['claim'] for x in t1)

    if t1_total_claims <= total_exit:
        for x in t1:
            x['payout'] = x['claim']
    else:
        ratio = total_exit / t1_total_claims if t1_total_claims > 0 else 0
        for x in t1:
            x['payout'] = x['claim'] * ratio

    t1_payout = sum(x['payout'] for x in t1)
    remaining1 = max(0.0, total_exit - t1_payout)

    # ── Tier 2: Seniority groups, pari passu within group ────────────────────
    t2 = []
    for pc in pref_classes_norm:
        claim = pc['capital_invested'] * pc['pref_multiple'] + pc['cumulative_dividends']
        t2.append({**pc, 'pref_claim': claim, 'pref_payout': 0.0,
                   'participation_payout': 0.0, 'converted': False, 'total_payout': 0.0})

    for level in sorted(set(p['seniority'] for p in t2)):
        group = [p for p in t2 if p['seniority'] == level]
        group_total = sum(p['pref_claim'] for p in group)
        avail = sum(p['pref_payout'] for p in t2 if p['seniority'] < level)
        avail = remaining1 - avail  # remaining after more-senior groups

        if group_total <= avail:
            for p in group:
                p['pref_payout'] = p['pref_claim']
        else:
            ratio = avail / group_total if group_total > 0 else 0
            for p in group:
                p['pref_payout'] = p['pref_claim'] * ratio

    t2_pref_total = sum(p['pref_payout'] for p in t2)
    remaining2 = max(0.0, remaining1 - t2_pref_total)

    # ── Conversion Check: non-participating preferred ─────────────────────────
    # Each non-participating class takes MAX(liq_pref_payout, as_converted_value).
    # Iterative: converters join the common pool, freeing their T2 payout for redistribution.
    converting_names = set()

    for _ in range(len(t2) + 1):
        # Shares already in common pool from previously-identified converters
        conv_shares = sum(p['shares'] for p in t2 if p['name'] in converting_names)

        # T2 payout from non-converters only
        non_conv_t2 = sum(
            p['pref_payout'] for p in t2 if p['name'] not in converting_names
        )

        # Remaining for the common pool when converters give back their T2 payout
        pool = remaining1 - non_conv_t2
        total_pool_shares = common_shares + conv_shares
        pps = pool / total_pool_shares if total_pool_shares > 0 else 0.0

        newly_converting = set()
        for p in t2:
            if p['name'] in converting_names:
                continue
            if p['participation'] != 'non-participating':
                continue  # participating preferred is handled differently
            as_converted = p['shares'] * pps
            if as_converted > p['pref_payout'] + 1e-2:
                newly_converting.add(p['name'])

        if not newly_converting:
            break
        converting_names.update(newly_converting)

    # Apply conversion results
    for p in t2:
        if p['name'] in converting_names:
            p['converted'] = True
            p['pref_payout'] = 0.0

    # Recompute after conversion
    t2_pref_total = sum(p['pref_payout'] for p in t2)
    remaining2 = max(0.0, remaining1 - t2_pref_total)

    # ── Tier 3: Common pool + participating + converted ───────────────────────
    participating = [p for p in t2 if p['participation'] != 'non-participating' and not p['converted']]
    converted = [p for p in t2 if p['converted']]

    base_pool_shares = common_shares + sum(p['shares'] for p in converted)

    if not participating:
        pps = remaining2 / base_pool_shares if base_pool_shares > 0 else 0.0
        for p in t2:
            p['participation_payout'] = 0.0
    else:
        # Iterative solve for capped participation
        uncapped = list(participating)
        capped_amounts = {}

        for _ in range(len(participating) + 1):
            part_shares = sum(p['shares'] for p in uncapped)
            total_pool = base_pool_shares + part_shares
            pool_avail = remaining2 - sum(capped_amounts.values())
            pps = pool_avail / total_pool if total_pool > 0 else 0.0

            newly_capped = []
            for p in uncapped:
                if p['participation'] == 'capped' and p['participation_cap']:
                    max_total = p['participation_cap'] * p['capital_invested'] + p['cumulative_dividends']
                    part_amt = p['shares'] * pps
                    if p['pref_payout'] + part_amt > max_total:
                        capped_amounts[p['name']] = max(0.0, max_total - p['pref_payout'])
                        newly_capped.append(p)
            if not newly_capped:
                break
            uncapped = [p for p in uncapped if p not in newly_capped]

        for p in t2:
            if p['participation'] == 'non-participating':
                p['participation_payout'] = 0.0
            elif p['converted']:
                p['participation_payout'] = 0.0
            elif p['name'] in capped_amounts:
                p['participation_payout'] = capped_amounts[p['name']]
            else:
                p['participation_payout'] = p['shares'] * pps

    # ── SAFE conversion check (after pps is established) ─────────────────────
    # If a conv instrument has implied_shares > 0, it takes MAX(claim, shares × pps)
    for x in t1:
        if x['implied_shares'] > 0:
            equity_val = x['implied_shares'] * pps
            if equity_val > x['payout']:
                x['payout'] = equity_val
                x['converted_to_equity'] = True
            else:
                x['converted_to_equity'] = False

    # Final per-class totals
    common_total = base_pool_shares * pps
    for p in t2:
        if p['converted']:
            p['total_payout'] = p['shares'] * pps
        else:
            p['total_payout'] = p['pref_payout'] + p['participation_payout']

    return {
        'tier1_payout': t1_payout,
        'tier2_pref_total': t2_pref_total,
        'remaining1': remaining1,
        'remaining2': remaining2,
        'pps': pps,
        'common_total': common_total,
        'tier1_instruments': t1,
        'tier2_prefs': t2,
        'has_participating': bool(participating),
        'has_converted': bool(converted),
        'converting_names': converting_names,
        'capped_classes': set(capped_amounts.keys()) if participating else set(),
    }


# ─── Sheet 1: Waterfall Analysis ──────────────────────────────────────────────

def build_waterfall_sheet(wb, company_name, cap_table_date,
                           total_exit, pref_multiple,
                           conv_instruments,
                           pref_classes,
                           common_shares,
                           total_options=0,
                           total_warrants=0):
    """
    Build Sheet 1 (Waterfall Analysis).

    Parameters
    ----------
    pref_multiple : float
        Default preference multiple used for any class that doesn't specify its own.

    conv_instruments : list of tuple or dict
        Old: (name, face_value)
        New: see _normalize_conv() docstring

    pref_classes : list of tuple or dict
        Old: (name, shares, capital_invested) or (name, shares, capital_invested, multiple)
        New: see _normalize_pref() docstring

    Returns dict with refs for cross-sheet formula references.
    """
    cn = [_normalize_conv(c) for c in conv_instruments]
    pn = [_normalize_pref(c, pref_multiple) for c in pref_classes]

    wf = compute_waterfall(total_exit, cn, pn, common_shares)

    has_participating = wf['has_participating']
    has_converted     = wf['has_converted']
    has_capped        = bool(wf['capped_classes'])
    has_dividends     = any(p['cumulative_dividends'] > 0 for p in pn)
    has_multi_seniority = len(set(p['seniority'] for p in pn)) > 1
    has_accruals      = any(c['accrued_interest'] > 0 or c['coc_multiple'] != 1.0 for c in cn)
    has_safe_conv     = any(c['implied_shares'] > 0 for c in cn)

    ws = wb.active
    ws.title = 'Waterfall Analysis'
    ws.sheet_view.showGridLines = False

    # Columns: A=label, B=shares, C=capital/principal, D=div/interest, E=multiple, F=claim, G=payout, H=notes
    NCOLS = 8
    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 26

    refs = {}
    r = 1

    exit_fmt = f'${total_exit/1e6:.0f}M' if total_exit % 1e6 == 0 else f'${total_exit:,.0f}'
    pref_desc = 'Multiple Tiers' if len(set(p['pref_multiple'] for p in pn)) > 1 else f'{pref_multiple:.0f}x Liquidation Preference'

    # Title
    section_header(ws, r, NCOLS, f'{company_name}  ·  Liquidation Waterfall Analysis', C['nav'], height=26)
    ws.cell(r, 1).font = Font(name='Arial', bold=True, color=C['white'], size=13)
    r += 1
    section_header(ws, r, NCOLS,
        f'Exit Proceeds: {exit_fmt}  |  {pref_desc}  |  Cap Table as of {cap_table_date}',
        C['subtitle'], fc='404040')
    ws.cell(r, 1).font = Font(name='Arial', italic=True, color='404040', size=10)
    r += 2

    # ── Assumptions Block ─────────────────────────────────────────────────────
    section_header(ws, r, NCOLS, 'ASSUMPTIONS', C['blue_hdr'])
    r += 1

    h(ws, r, 1, 'Total Exit Proceeds', bold=True)
    h(ws, r, 2, total_exit, fc=C['blue_input'], bg=C['input_bg'], num_fmt=CURRENCY)
    h(ws, r, 3, '← blue cells = hardcoded inputs', italic=True, fc=C['gray_text'], size=9)
    refs['EXIT_R'], refs['EXIT_C'] = r, 2
    EXIT_REF = f'$B${r}'
    r += 1

    h(ws, r, 1, 'Default Preference Multiple', bold=True)
    h(ws, r, 2, pref_multiple, fc=C['blue_input'], bg=C['input_bg'], num_fmt=MULT)
    h(ws, r, 3, 'Per-class multiples set in Tier 2 below' if len(set(p['pref_multiple'] for p in pn)) > 1 else '', italic=True, fc=C['gray_text'], size=9)
    refs['LPM_R'], refs['LPM_C'] = r, 2
    r += 1

    h(ws, r, 1, 'Preferred Basis', bold=True)
    h(ws, r, 2, 'Capital Invested + Cumulative Dividends' if has_dividends else 'Capital Invested', fc=C['blue_input'], bg=C['input_bg'])
    r += 1

    if has_converted:
        h(ws, r, 1, 'Conversion Check', bold=True)
        h(ws, r, 2, 'Preferred converts when as-converted > liq pref (see Tier 2)', fc=C['blue_input'], bg=C['input_bg'])
        h(ws, r, 3, f'{len(wf["converting_names"])} class(es) converted in this scenario', italic=True, fc=C['gray_text'], size=9)
        r += 1

    if has_multi_seniority:
        h(ws, r, 1, 'Seniority', bold=True)
        levels = sorted(set(p['seniority'] for p in pn))
        groups = [f'S{lv}: ' + ', '.join(p['name'] for p in pn if p['seniority'] == lv) for lv in levels]
        h(ws, r, 2, '  |  '.join(groups), fc=C['blue_input'], bg=C['input_bg'], wrap=True)
        rh(ws, r, 28)
        r += 1

    r += 1  # spacer

    # ── Tier 1: Convertibles / SAFEs ──────────────────────────────────────────
    section_header(ws, r, NCOLS,
        '  TIER 1  ·  CONVERTIBLE NOTES & SAFEs  ·  Pari Passu', C['green_hdr'])
    r += 1

    t1_cols = ['Instrument', 'Principal', 'Accrued Int / CoC Premium', '',
               '', 'Total Claim', 'Payout', 'Notes']
    col_headers(ws, r, t1_cols, '548235', row_height=28)
    r += 1

    t1_rows = []
    for idx, c in enumerate(cn):
        inst = wf['tier1_instruments'][idx]
        has_note = c['accrued_interest'] > 0 or c['coc_multiple'] != 1.0 or c.get('converted_to_equity')
        note_text = ''
        if c['coc_multiple'] != 1.0:
            note_text += f'{c["coc_multiple"]:.1f}x CoC multiple  '
        if c['accrued_interest'] > 0:
            note_text += f'+${c["accrued_interest"]:,.0f} accrued int  '
        if c.get('implied_shares', 0) > 0:
            eq_val = c['implied_shares'] * wf['pps']
            if inst.get('converted_to_equity'):
                note_text += f'Converted to equity (${eq_val:,.0f} > ${inst["claim"]:,.0f} claim) ⁴'
            else:
                note_text += f'Implied equity ${eq_val:,.0f} < claim — not converting'

        h(ws, r, 1, f'  {c["name"]}', bg=C['green_bg'])
        h(ws, r, 2, c['principal'], fc=C['blue_input'], bg=C['green_bg'], num_fmt=CURRENCY, align='right')
        h(ws, r, 3, c['accrued_interest'] + (c['principal'] * (c['coc_multiple'] - 1.0)) if (c['accrued_interest'] > 0 or c['coc_multiple'] != 1.0) else '—',
          fc=C['blue_input'] if has_note else C['gray_text'],
          bg=C['green_bg'], num_fmt=CURRENCY if has_note else None, align='right')
        h(ws, r, 6, inst['claim'], bg=C['green_bg'], num_fmt=CURRENCY, align='right')
        h(ws, r, 7, inst['payout'], bold=True, bg=C['green_bg'], num_fmt=CURRENCY, align='right')
        if note_text:
            h(ws, r, 8, note_text.strip(), italic=True, fc=C['gray_text'], size=8, bg=C['green_bg'], wrap=True)
            rh(ws, r, 28)
        t1_rows.append(r)
        r += 1

    CONV_S, CONV_E = t1_rows[0], t1_rows[-1]
    h(ws, r, 1, '  Total Convertibles', bold=True, bg=C['green_sub'])
    h(ws, r, 6, sum(x['claim'] for x in wf['tier1_instruments']),
      bold=True, bg=C['green_sub'], num_fmt=CURRENCY, align='right')
    h(ws, r, 7, wf['tier1_payout'], bold=True, bg=C['green_sub'], num_fmt=CURRENCY, align='right')
    CONV_TOT_R = r
    r += 1

    h(ws, r, 1, '  Remaining After Tier 1', bold=True)
    h(ws, r, 7, wf['remaining1'], bold=True, bg=C['green_bg'], num_fmt=CURRENCY, align='right')
    REM1_R = r
    r += 2

    # ── Tier 2: Preferred ─────────────────────────────────────────────────────
    t2_hdr = '  TIER 2  ·  PREFERRED STOCK  ·  Seniority Groups; Pari Passu Within Group'
    if has_converted:
        t2_hdr += '  ·  Conversion Check Applied'
    section_header(ws, r, NCOLS, t2_hdr, C['blue_hdr'])
    r += 1

    t2_cols = ['Class', 'Shares', 'Capital Invested',
               'Cumulative Dividends', 'Pref Multiple',
               'Liq Pref Claim', 'T2 Payout', 'Structure / Conversion']
    col_headers(ws, r, t2_cols, '2E74B5', row_height=28)
    r += 1

    pref_rows = {}         # name → row (for cross-sheet refs)
    t2_data = wf['tier2_prefs']
    current_seniority = None

    pref_sec_s = r
    for p in t2_data:
        # Seniority sub-header
        if p['seniority'] != current_seniority:
            current_seniority = p['seniority']
            seniority_group = [x for x in t2_data if x['seniority'] == current_seniority]
            group_label = (
                f'  Seniority {current_seniority} — Pari Passu'
                + (' (Most Senior)' if current_seniority == min(p['seniority'] for p in t2_data) else '')
            )
            section_header(ws, r, NCOLS, group_label, '4472C4', height=18)
            r += 1

        converted = p['converted']
        bg = C['converted_bg'] if converted else C['blue_bg']
        fc_val = C['converted_text'] if converted else C['blue_input']

        h(ws, r, 1, f'  {p["name"]}', bg=bg, fc=C['converted_text'] if converted else C['black'])
        h(ws, r, 2, p['shares'], fc=fc_val, bg=bg, num_fmt=NUMFMT, align='right')
        h(ws, r, 3, p['capital_invested'], fc=fc_val, bg=bg, num_fmt=CURRENCY2, align='right')
        h(ws, r, 4, p['cumulative_dividends'] if p['cumulative_dividends'] > 0 else '—',
          fc=fc_val, bg=bg,
          num_fmt=CURRENCY if p['cumulative_dividends'] > 0 else None,
          align='right')
        h(ws, r, 5, p['pref_multiple'], fc=fc_val, bg=bg, num_fmt=MULT, align='center')
        h(ws, r, 6, p['pref_claim'], bg=bg, num_fmt=CURRENCY, align='right',
          fc=C['converted_text'] if converted else C['black'])

        if converted:
            h(ws, r, 7, 0.0, bg=bg, num_fmt=CURRENCY, align='right', fc=C['converted_text'])
            as_conv = p['shares'] * wf['pps']
            h(ws, r, 8, f'Converted to common ⁵  (as-converted ${as_conv:,.0f} > liq pref ${p["pref_claim"]:,.0f})',
              italic=True, fc=C['converted_text'], size=8, bg=bg, wrap=True)
            rh(ws, r, 28)
        else:
            h(ws, r, 7, p['pref_payout'], bold=True, bg=bg, num_fmt=CURRENCY, align='right')
            label = _participation_label(p)
            if p['pref_payout'] < p['pref_claim'] - 1:
                label += '  ⚠ Pro-rated (insufficient funds)'
            h(ws, r, 8, label, italic=True, fc=C['gray_text'], size=9, bg=bg)

        pref_rows[p['name']] = r
        refs['PREF_R'] = refs.get('PREF_R', r)  # first class for backward compat
        r += 1

    pref_sec_e = r - 1

    h(ws, r, 1, '  Total Preferred (Preference Payout)', bold=True, bg=C['blue_sub'])
    h(ws, r, 7, wf['tier2_pref_total'], bold=True, bg=C['blue_sub'], num_fmt=CURRENCY, align='right')
    PREF_TOT_R = r
    r += 1

    h(ws, r, 1, '  Remaining After Tier 2 Preference', bold=True)
    h(ws, r, 7, wf['remaining2'], bold=True, bg=C['blue_bg'], num_fmt=CURRENCY, align='right')
    refs['REM2_R'], refs['REM2_C'] = r, 7
    r += 2

    # ── Tier 3: Common pool ───────────────────────────────────────────────────
    tier3_hdr = '  TIER 3  ·  COMMON POOL  ·  Pro-Rata Residual'
    if has_converted:
        tier3_hdr += '  (includes converted preferred)'
    if has_participating:
        tier3_hdr += '  (includes participating preferred)'
    section_header(ws, r, NCOLS, tier3_hdr, C['gold_hdr'])
    r += 1
    col_headers(ws, r, ['Class', 'Pool Shares', '', '', '', '',
                         'Per-Share Value', 'Payout'], 'BF8F00', row_height=22)
    r += 1

    # Denominator helper row
    pool_shares_norm = common_shares
    if has_converted:
        pool_shares_norm += sum(p['shares'] for p in t2_data if p['converted'])
    if has_participating:
        pool_shares_norm += sum(p['shares'] for p in t2_data
                                if p['participation'] != 'non-participating' and not p['converted']
                                and p['name'] not in wf['capped_classes'])

    h(ws, r, 1, '  Common Pool Denominator (participating shares)', bold=True, fc=C['gray_text'], italic=True)
    h(ws, r, 2, pool_shares_norm, bold=True, bg=C['input_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
    DENOM_ROW = r
    r += 1

    h(ws, r, 1, '  Per-Share Value', bold=True)
    h(ws, r, 7, wf['pps'], bold=True, bg=C['gold_bg'], num_fmt=CURRENCY2, align='right')
    refs['PPS_ROW'] = r
    PPS_ROW = r
    r += 1

    # Common shares row
    h(ws, r, 1, '  Common Stock', bg=C['gold_bg'])
    h(ws, r, 2, common_shares, fc=C['blue_input'], bg=C['gold_bg'], num_fmt=NUMFMT, align='right')
    h(ws, r, 8, wf['common_total'] - sum(p['shares'] * wf['pps'] for p in t2_data if p['converted']),
      bold=True, bg=C['gold_sub'], num_fmt=CURRENCY, align='right')
    refs['COMM_R'] = r
    r += 1

    # Converted preferred rows
    if has_converted:
        for p in t2_data:
            if not p['converted']:
                continue
            conv_payout = p['shares'] * wf['pps']
            h(ws, r, 1, f'  {p["name"]} — Converted ⁵', bg=C['gold_bg'], italic=True)
            h(ws, r, 2, p['shares'], bg=C['gold_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
            h(ws, r, 8, conv_payout, bold=True, bg=C['gold_sub'], num_fmt=CURRENCY, align='right')
            r += 1

    # Participating preferred rows
    if has_participating:
        section_header(ws, r, NCOLS, '  PREFERRED PARTICIPATION (in Common Residual)', '4472C4', height=18)
        r += 1
        for p in t2_data:
            if p['participation'] == 'non-participating' or p['converted']:
                continue
            is_capped = p['name'] in wf['capped_classes']
            h(ws, r, 1, f'  {p["name"]} — Participation{"  (CAPPED ³)" if is_capped else ""}',
              bg=C['blue_bg'], italic=is_capped)
            h(ws, r, 2, p['shares'], bg=C['blue_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
            h(ws, r, 8, p['participation_payout'],
              bold=True, bg=C['blue_bg'], num_fmt=CURRENCY, align='right',
              fc=C['blue_input'] if is_capped else C['black'])
            if is_capped:
                cap_val = p['participation_cap'] * p['capital_invested'] + p['cumulative_dividends']
                h(ws, r, NCOLS, f'Cap: {p["participation_cap"]:.1f}x invested → max ${cap_val:,.0f} total',
                  italic=True, fc=C['gray_text'], size=8)
            r += 1

    # Tier 3 total
    t3_total = wf['common_total'] + sum(p['participation_payout'] for p in t2_data)
    h(ws, r, 1, '  Total Tier 3', bold=True, bg=C['gold_sub'])
    h(ws, r, 8, t3_total, bold=True, bg=C['gold_sub'], num_fmt=CURRENCY, align='right')
    r += 2

    # ── Waterfall Summary ─────────────────────────────────────────────────────
    section_header(ws, r, NCOLS, 'WATERFALL SUMMARY', C['nav'], height=22)
    ws.cell(r, 1).font = Font(name='Arial', bold=True, color=C['white'], size=11)
    r += 1
    col_headers(ws, r,
        ['Stakeholder Group', 'Shares / Principal', 'Liq Pref Claim', 'T2 Pref Payout',
         'T3 Participation', 'Total Payout', '% of Total', ''],
        C['nav'], row_height=28)
    r += 1

    # Convertibles row
    h(ws, r, 1, '  Convertible Notes & SAFEs', bold=True, bg=C['green_bg'])
    h(ws, r, 2, sum(c['principal'] for c in cn), bg=C['green_bg'], num_fmt=CURRENCY, align='right')
    h(ws, r, 3, '1x Principal + Accruals', bg=C['green_bg'], align='center', fc=C['gray_text'], size=9, italic=True)
    h(ws, r, 4, wf['tier1_payout'], bold=True, bg=C['green_bg'], num_fmt=CURRENCY, align='right')
    h(ws, r, 5, '—', bg=C['green_bg'], align='center', fc=C['gray_text'])
    h(ws, r, 6, wf['tier1_payout'], bold=True, bg=C['green_bg'], num_fmt=CURRENCY, align='right')
    h(ws, r, 7, formula=f'=F{r}/{EXIT_REF}', bg=C['green_bg'], num_fmt=PCT, align='center')
    r += 1

    # Preferred rows (one per class)
    for p in t2_data:
        bg = C['converted_bg'] if p['converted'] else C['blue_bg']
        fc_t = C['converted_text'] if p['converted'] else C['black']
        h(ws, r, 1, f'  {p["name"]}', bold=True, bg=bg, fc=fc_t)
        h(ws, r, 2, p['shares'], bg=bg, num_fmt=NUMFMT, align='right', fc=fc_t)
        h(ws, r, 3, p['pref_claim'], bg=bg, num_fmt=CURRENCY, align='right', fc=fc_t)
        h(ws, r, 4, p['pref_payout'], bold=True, bg=bg, num_fmt=CURRENCY, align='right', fc=fc_t)
        t3_part = p['total_payout'] - p['pref_payout']
        h(ws, r, 5, t3_part if t3_part > 0 else '—',
          bg=bg, num_fmt=CURRENCY if t3_part > 0 else None, align='right' if t3_part > 0 else 'center',
          fc=fc_t if t3_part > 0 else C['gray_text'])
        h(ws, r, 6, p['total_payout'], bold=True, bg=bg, num_fmt=CURRENCY, align='right', fc=fc_t)
        h(ws, r, 7, formula=f'=F{r}/{EXIT_REF}', bg=bg, num_fmt=PCT, align='center', fc=fc_t)
        if p['converted']:
            h(ws, r, 8, 'Converted to common ⁵', italic=True, fc=C['converted_text'], size=8)
        r += 1

    # Common row
    common_payout = wf['common_total']
    h(ws, r, 1, '  Common Stock', bold=True, bg=C['gold_bg'])
    h(ws, r, 2, common_shares, bg=C['gold_bg'], num_fmt=NUMFMT, align='right')
    h(ws, r, 3, 'Pro-Rata Residual', bg=C['gold_bg'], align='center', fc=C['gray_text'], size=9, italic=True)
    h(ws, r, 4, '—', bg=C['gold_bg'], align='center', fc=C['gray_text'])
    h(ws, r, 5, '—', bg=C['gold_bg'], align='center', fc=C['gray_text'])
    h(ws, r, 6, common_payout, bold=True, bg=C['gold_bg'], num_fmt=CURRENCY, align='right')
    h(ws, r, 7, formula=f'=F{r}/{EXIT_REF}', bg=C['gold_bg'], num_fmt=PCT, align='center')
    COMM_SUM_R = r
    r += 1

    if total_warrants:
        h(ws, r, 1, '  Warrants — Not Participating ¹', bg=C['gray_bg'], fc=C['gray_text'])
        h(ws, r, 2, total_warrants, bg=C['gray_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
        for col in [4, 5, 6]: h(ws, r, col, 0, bg=C['gray_bg'], num_fmt=CURRENCY, align='right', fc=C['gray_text'])
        h(ws, r, 7, '—', bg=C['gray_bg'], align='center', fc=C['gray_text'])
        r += 1

    if total_options:
        h(ws, r, 1, '  Options / RSUs (Theoretical Gross) ²', bg=C['gray_bg'], fc=C['gray_text'])
        h(ws, r, 2, total_options, bg=C['gray_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
        opt_gross = total_options * wf['pps']
        h(ws, r, 6, opt_gross, bg=C['gray_bg'], num_fmt=CURRENCY, align='right', fc=C['gray_text'])
        h(ws, r, 7, '—', bg=C['gray_bg'], align='center', fc=C['gray_text'])
        r += 1

    # Grand total
    grand_total = wf['tier1_payout'] + wf['tier2_pref_total'] + t3_total
    h(ws, r, 1, '  TOTAL EXIT PROCEEDS', bold=True, bg=C['total_bg'])
    h(ws, r, 6, grand_total, bold=True, bg=C['total_bg'], num_fmt=CURRENCY, align='right')
    h(ws, r, 7, formula=f'=F{r}/{EXIT_REF}', bold=True, bg=C['total_bg'], num_fmt=PCT, align='center')
    r += 2

    # Footnotes
    notes = []
    if total_warrants:
        notes.append('¹ Warrants excluded — exercise price not in cap table.')
    if total_options:
        notes.append('² Options shown GROSS. Net = (per-share value − exercise price) × options.')
    if has_capped:
        notes.append('³ Capped participation computed via iterative Python solve (compute_waterfall()). '
                     'Values shown in blue are pre-computed and written as inputs.')
    if has_safe_conv:
        notes.append('⁴ SAFE/note conversion check: instrument takes MAX(principal claim, implied_shares × common PPS).')
    if has_converted:
        notes.append('⁵ Converted preferred: class takes as-converted common value (MAX of liq pref vs. as-converted). '
                     'Conversion check is iterative — converting classes join the common pool and give back their T2 payout.')

    for note in notes:
        h(ws, r, 1, note, italic=True, fc=C['gray_text'], size=8, wrap=True)
        ws.row_dimensions[r].height = 24
        r += 1

    refs['wf'] = wf
    return refs


# ─── Sheet 2: Individual Payouts ──────────────────────────────────────────────

def build_individual_payouts_sheet(wb, company_name, cap_table_date,
                                    total_exit, refs,
                                    named_individuals,
                                    all_conv_instruments,
                                    all_pref_holders,
                                    all_common_holders,
                                    all_option_holders=None,
                                    all_warrant_holders=None):
    """
    Build Sheet 2 (Individual Payouts).

    named_individuals : list of (str, int, int, int)
        (name, common_shares, pref_shares, options)
    all_pref_holders : list of (str, int, bool)
        (holder_name, pref_shares, is_named)
    all_common_holders : list of (str, int, bool)
        (holder_name, common_shares, is_named)
    """
    ws = wb.create_sheet('Individual Payouts')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14

    PPS_ROW = refs.get('PPS_ROW', refs.get('COMM_R'))
    CPPS_REF = f"='Waterfall Analysis'!G{PPS_ROW}"
    PPPS_REF = f"='Waterfall Analysis'!G{refs['PREF_R']}"
    exit_fmt = f'${total_exit/1e6:.0f}M' if total_exit % 1e6 == 0 else f'${total_exit:,.0f}'
    EXIT_VAL = total_exit

    r = 1
    section_header(ws, r, 6, f'{company_name}  ·  Payout Breakdown by Stakeholder', C['nav'], height=26)
    ws.cell(r, 1).font = Font(name='Arial', bold=True, color=C['white'], size=13)
    r += 1
    section_header(ws, r, 6, f'{exit_fmt} Exit  |  Cap Table as of {cap_table_date}', C['subtitle'], fc='404040')
    ws.cell(r, 1).font = Font(name='Arial', italic=True, color='404040', size=10)
    r += 2

    section_header(ws, r, 6, 'KEY ASSUMPTIONS (linked to Waterfall Analysis)', C['blue_hdr'], height=20)
    r += 1
    h(ws, r, 1, 'Common stock per-share value', bold=True)
    h(ws, r, 2, formula=CPPS_REF, bg=C['input_bg'], num_fmt=CURRENCY2, align='right', fc='008000')
    CPPS = f'B{r}'
    r += 1
    h(ws, r, 1, 'Preferred T2 payout per-share (1st class)', bold=True)
    h(ws, r, 2, formula=PPPS_REF, bg=C['input_bg'], num_fmt=CURRENCY2, align='right', fc='008000')
    PPPS = f'B{r}'
    r += 2

    # Named individual summary
    section_header(ws, r, 6, 'NAMED INDIVIDUAL SUMMARY', C['highlight_hdr2'], height=22)
    r += 1
    col_headers(ws, r,
        ['Stakeholder', 'Common Payout', 'Preferred Payout', 'Option Payout (Gross) ²', 'TOTAL PAYOUT', '% of Exit'],
        C['highlight_hdr'], row_height=28)
    r += 1

    named_s = r
    for name, cs, ps, opts in named_individuals:
        cp = f'={cs}*{CPPS}' if cs else 0
        pp = f'={ps}*{PPPS}' if ps else 0
        op = f'={opts}*{CPPS}' if opts else 0
        h(ws, r, 1, f'  {name}', bold=True, bg=C['highlight'], wrap=True)
        h(ws, r, 2, formula=cp if isinstance(cp, str) else None, value=cp if isinstance(cp, int) else None,
          bg=C['highlight'], num_fmt=CURRENCY, align='right')
        h(ws, r, 3, formula=pp if isinstance(pp, str) else None, value=pp if isinstance(pp, int) else None,
          bg=C['highlight'], num_fmt=CURRENCY, align='right')
        h(ws, r, 4, formula=op if isinstance(op, str) else None, value=op if isinstance(op, int) else None,
          bg=C['highlight'], num_fmt=CURRENCY, align='right', fc=C['gray_text'])
        h(ws, r, 5, formula=f'=B{r}+C{r}+D{r}', bold=True, bg=C['highlight'], num_fmt=CURRENCY, align='right')
        h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bg=C['highlight'], num_fmt=PCT, align='center')
        rh(ws, r, 32 if '\n' in str(name) else 18)
        r += 1

    h(ws, r, 1, '  Subtotal — Named Individuals', bold=True, bg=C['total_bg'])
    for col_i in range(2, 7):
        cl = get_column_letter(col_i)
        h(ws, r, col_i, formula=f'=SUM({cl}{named_s}:{cl}{r-1})', bold=True,
          bg=C['total_bg'], num_fmt=CURRENCY if col_i < 6 else PCT, align='right')
    r += 2

    # Full breakdown
    section_header(ws, r, 6, 'FULL PAYOUT BREAKDOWN BY STAKEHOLDER GROUP', C['nav'], height=22)
    r += 1
    col_headers(ws, r,
        ['Stakeholder / Group', 'Common Shares', 'Pref Shares', 'Options', 'TOTAL PAYOUT', '% of Exit'],
        C['nav'], row_height=28)
    r += 1

    # Tier 1
    section_header(ws, r, 6, '  TIER 1 · CONVERTIBLES — Pari Passu', C['green_hdr'], height=20)
    r += 1
    conv_s = r
    for name, amt in all_conv_instruments:
        h(ws, r, 1, f'    {name}', bg=C['green_bg'])
        for c in [2, 3, 4]: h(ws, r, c, '—', bg=C['green_bg'], align='center', fc=C['gray_text'])
        h(ws, r, 5, amt, bold=True, bg=C['green_bg'], num_fmt=CURRENCY, align='right')
        h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bg=C['green_bg'], num_fmt=PCT, align='center')
        r += 1
    conv_e = r - 1
    h(ws, r, 1, '  Total Convertibles', bold=True, bg=C['green_sub'])
    for c in [2, 3, 4]: h(ws, r, c, '', bg=C['green_sub'])
    h(ws, r, 5, formula=f'=SUM(E{conv_s}:E{conv_e})', bold=True, bg=C['green_sub'], num_fmt=CURRENCY, align='right')
    h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bold=True, bg=C['green_sub'], num_fmt=PCT, align='center')
    r += 2

    # Tier 2
    section_header(ws, r, 6, '  TIER 2 · PREFERRED — Preference Payout', C['blue_hdr'], height=20)
    r += 1
    pref_s = r
    for name, shares, is_named in all_pref_holders:
        bg = C['highlight'] if is_named else C['blue_bg']
        h(ws, r, 1, f'    {name}', bg=bg, bold=is_named)
        h(ws, r, 2, '—', bg=bg, align='center', fc=C['gray_text'])
        h(ws, r, 3, shares, bg=bg, num_fmt=NUMFMT, align='right')
        h(ws, r, 4, '—', bg=bg, align='center', fc=C['gray_text'])
        h(ws, r, 5, formula=f'={shares}*{PPPS}', bg=bg, num_fmt=CURRENCY, align='right', bold=is_named)
        h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bg=bg, num_fmt=PCT, align='center')
        r += 1
    pref_e = r - 1
    h(ws, r, 1, '  Total Preferred', bold=True, bg=C['blue_sub'])
    h(ws, r, 3, formula=f'=SUM(C{pref_s}:C{pref_e})', bold=True, bg=C['blue_sub'], num_fmt=NUMFMT, align='right')
    h(ws, r, 5, formula=f'=SUM(E{pref_s}:E{pref_e})', bold=True, bg=C['blue_sub'], num_fmt=CURRENCY, align='right')
    h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bold=True, bg=C['blue_sub'], num_fmt=PCT, align='center')
    r += 2

    # Tier 3
    section_header(ws, r, 6, '  TIER 3 · COMMON STOCK — Pro-Rata Residual', C['gold_hdr'], height=20)
    r += 1
    comm_s = r
    for name, shares, is_named in all_common_holders:
        bg = C['highlight'] if is_named else C['gold_bg']
        h(ws, r, 1, f'    {name}', bg=bg, bold=is_named)
        h(ws, r, 2, shares, bg=bg, num_fmt=NUMFMT, align='right')
        h(ws, r, 3, '—', bg=bg, align='center', fc=C['gray_text'])
        h(ws, r, 4, '—', bg=bg, align='center', fc=C['gray_text'])
        h(ws, r, 5, formula=f'={shares}*{CPPS}', bg=bg, num_fmt=CURRENCY, align='right', bold=is_named)
        h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bg=bg, num_fmt=PCT, align='center')
        r += 1
    comm_e = r - 1
    h(ws, r, 1, '  Total Common', bold=True, bg=C['gold_sub'])
    h(ws, r, 2, formula=f'=SUM(B{comm_s}:B{comm_e})', bold=True, bg=C['gold_sub'], num_fmt=NUMFMT, align='right')
    h(ws, r, 5, formula=f'=SUM(E{comm_s}:E{comm_e})', bold=True, bg=C['gold_sub'], num_fmt=CURRENCY, align='right')
    h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bold=True, bg=C['gold_sub'], num_fmt=PCT, align='center')
    r += 2

    if all_warrant_holders:
        section_header(ws, r, 6, '  WARRANTS — Not Participating ¹', C['gray_hdr'], height=20)
        r += 1
        for name, shares in all_warrant_holders:
            h(ws, r, 1, f'    {name}', bg=C['gray_bg'], fc=C['gray_text'])
            h(ws, r, 2, shares, bg=C['gray_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
            h(ws, r, 5, 0, bg=C['gray_bg'], num_fmt=CURRENCY, align='right', fc=C['gray_text'])
            h(ws, r, 6, '—', bg=C['gray_bg'], align='center', fc=C['gray_text'])
            r += 1
        r += 1

    if all_option_holders:
        section_header(ws, r, 6, '  OPTIONS & RSUs — Theoretical Gross ²', C['gray_hdr'], height=20)
        r += 1
        col_headers(ws, r, ['Holder', '—', '—', 'Options / RSUs', 'Gross Payout ²', '% of Exit'],
                    C['gray_hdr'], row_height=18)
        r += 1
        opt_s = r
        for name, opts, is_named in all_option_holders:
            bg = C['highlight'] if is_named else C['gray_bg']
            fc_t = C['black'] if is_named else C['gray_text']
            h(ws, r, 1, f'    {name}', bg=bg, bold=is_named, fc=fc_t)
            h(ws, r, 2, '—', bg=bg, align='center', fc=C['gray_text'])
            h(ws, r, 3, '—', bg=bg, align='center', fc=C['gray_text'])
            h(ws, r, 4, opts, bg=bg, num_fmt=NUMFMT, align='right', fc=fc_t)
            h(ws, r, 5, formula=f'={opts}*{CPPS}', bg=bg, num_fmt=CURRENCY, align='right', bold=is_named, fc=fc_t)
            h(ws, r, 6, formula=f'=E{r}/{EXIT_VAL}', bg=bg, num_fmt=PCT, align='center', fc=C['gray_text'])
            r += 1
        opt_e = r - 1
        h(ws, r, 1, '  Total Options / RSUs', bold=True, bg=C['gray_bg'], fc=C['gray_text'])
        h(ws, r, 4, formula=f'=SUM(D{opt_s}:D{opt_e})', bold=True, bg=C['gray_bg'], num_fmt=NUMFMT, align='right', fc=C['gray_text'])
        h(ws, r, 5, formula=f'=SUM(E{opt_s}:E{opt_e})', bold=True, bg=C['gray_bg'], num_fmt=CURRENCY, align='right', fc=C['gray_text'])
        r += 2

    if all_warrant_holders:
        h(ws, r, 1, '¹ Warrants excluded — exercise price unavailable.', italic=True, fc=C['gray_text'], size=8)
        r += 1
    if all_option_holders:
        h(ws, r, 1, '² Options shown GROSS. Net = (per-share − exercise price) × options.', italic=True, fc=C['gray_text'], size=8)